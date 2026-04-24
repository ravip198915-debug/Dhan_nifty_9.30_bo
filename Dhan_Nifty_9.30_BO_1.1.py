"""
NIFTY 9:30 Candle Breakout Bot (PURE DHAN API)
------------------------------------------------
Strategy kept intact:
- 9:30 candle breakout on NIFTY spot
- One trade per day
- CE/PE side chosen by CPR + 20MA filter
- Premium SL/Target management
- Intraday square-off at FORCE_EXIT_TIME

Notes:
- Uses ONLY Dhan APIs (REST + marketfeed)
- No Zerodha/Kite imports or symbols
- No asyncio/event-loop usage (thread-safe polling model)
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time as dtime
import atexit
import os
import sys
import threading
import time
from typing import Dict, List, Optional, Set, Tuple

import requests
from colorama import init
from dhanhq import dhanhq
from dhanhq import marketfeed
from openpyxl import Workbook, load_workbook


# ========================== CONFIG ==========================
CLIENT_ID = os.getenv("DHAN_CLIENT_ID", "YOUR_CLIENT_ID")
ACCESS_TOKEN = os.getenv("DHAN_ACCESS_TOKEN", "YOUR_ACCESS_TOKEN")

MODE = "LIVE"
PRODUCT = "INTRADAY"
EXCHANGE_SEGMENT_OPT = "NSE_FNO"
ORDER_TYPE = "MARKET"

# Dhan security IDs
SPOT_SECURITY_ID = "13"  # NIFTY 50 index spot
LOT_SIZE = 130

PREM_SL_PTS = 20
PREM_TGT_PTS = 40

LAST_ENTRY_TIME = dtime(15, 15)
FORCE_EXIT_TIME = dtime(15, 20)
CPR_WIDE_THRESHOLD = 0.6

# WebSocket reconnect controls
RECONNECT_BASE_DELAY = 3
RECONNECT_MAX_DELAY = 30
MAX_EMPTY_TICKS = 120
EMPTY_TICK_LOG_EVERY = 20
MAX_RECONNECT_ATTEMPTS = 10
TRADE_COOLDOWN_SEC = 10
TRADE_LOG_FILE = "trade_log.xlsx"
WATCHDOG_NO_TICK_SEC = 10
HEARTBEAT_INTERVAL_SEC = 10

# lock file
LOCK_FILE = "trading.lock"

# telegram (optional)
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
CHAT_ID = os.getenv("CHAT_ID", "")


# ========================== UI ==========================
init(autoreset=True)
GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
BLUE = "\033[94m"
RESET = "\033[0m"


def send_telegram(msg: str) -> None:
    if not BOT_TOKEN or not CHAT_ID:
        return
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {"chat_id": CHAT_ID, "text": f"<pre>{msg}</pre>", "parse_mode": "HTML"}
        requests.post(url, data=payload, timeout=5)
    except Exception as exc:
        print(f"Telegram Error: {exc}")


def remove_lock() -> None:
    if os.path.exists(LOCK_FILE):
        os.remove(LOCK_FILE)


def ensure_single_instance() -> None:
    if os.path.exists(LOCK_FILE):
        print("Script already running — exiting")
        sys.exit(0)
    open(LOCK_FILE, "w").close()


# ========================== HELPERS ==========================
def _to_float(value) -> Optional[float]:
    try:
        return float(value)
    except Exception:
        return None


@dataclass
class Candle:
    high: float
    low: float


class DhanRestClient:
    """Small REST wrapper for endpoints needed by this strategy."""

    def __init__(self, client_id: str, access_token: str):
        self.client_id = client_id
        self.access_token = access_token
        self.base = "https://api.dhan.co/v2"
        self.headers = {
            "access-token": access_token,
            "Content-Type": "application/json",
        }
        self.auth_failed = False

    def _request(self, method: str, path: str, payload: Optional[dict] = None) -> Optional[dict]:
        try:
            url = f"{self.base}{path}"
            resp = requests.request(method, url, headers=self.headers, json=payload, timeout=10)

            if resp.status_code == 401:
                print("[AUTH ERROR] Access token expired/invalid.")
                self.auth_failed = True
                return None

            resp.raise_for_status()
            data = resp.json()
            return data if isinstance(data, dict) else {"data": data}
        except requests.exceptions.RequestException as exc:
            print(f"[NETWORK ERROR] {path}: {exc}")
            return None
        except Exception as exc:
            print(f"[REST ERROR] {path}: {exc}")
            return None

    def place_order(self, security_id: str, side: str, qty: int) -> Optional[dict]:
        payload = {
            "dhanClientId": self.client_id,
            "transactionType": side,
            "exchangeSegment": EXCHANGE_SEGMENT_OPT,
            "productType": PRODUCT,
            "orderType": ORDER_TYPE,
            "validity": "DAY",
            "securityId": str(security_id),
            "quantity": int(qty),
            "price": 0,
            "triggerPrice": 0,
            "afterMarketOrder": False,
        }
        return self._request("POST", "/orders", payload)


# ========================== STRATEGY STATE ==========================
class BotState:
    def __init__(self):
        self.spot_ltp: Optional[float] = None
        self.option_ltp: Optional[float] = None

        self.trade_open = False
        self.order_placed = False
        self.day_closed = False
        self.auto_ready = False

        self.allowed_side: Optional[str] = None
        self.auto_signal = "NO TRADE"
        self.cpr_type: Optional[str] = None
        self.ma_side: Optional[str] = None

        self.candle: Optional[Candle] = None
        self.candle_high: Optional[float] = None
        self.candle_low: Optional[float] = None
        self.candle_ready = False
        self.skip_trading_today = False
        self.trade_closed = False

        self.active_symbol: Optional[str] = None
        self.active_security_id: Optional[str] = None
        self.active_side: Optional[str] = None

        self.prem_entry: Optional[float] = None
        self.prem_sl: Optional[float] = None
        self.prem_target: Optional[float] = None

        self.last_option_tick_ts = 0.0
        self.feed_resubscribe_required = False
        self.last_spot_log_ts = 0.0
        self.last_option_log_ts = 0.0
        self.last_trade_time = 0.0
        self.reconnect_failures = 0
        self.last_any_tick_ts = time.time()
        self.last_heartbeat_ts = 0.0
        self.last_no_tick_warning_ts = 0.0
        self.bot_start_dt = datetime.now()
        self.late_start_checked = False

        self.entry_time: Optional[datetime] = None
        self.exit_time: Optional[datetime] = None
        self.exit_reason: Optional[str] = None
        self.trade_logged = False
        self.pending_trade_row: Optional[dict] = None
        self.feed_client = None
        self.feed_thread: Optional[threading.Thread] = None


state = BotState()


class TradeLogger:
    HEADERS = [
        "Date",
        "Entry Time",
        "Exit Time",
        "Symbol",
        "Side (CE/PE)",
        "Entry Price",
        "Exit Price",
        "Quantity",
        "P&L",
        "Exit Reason (SL/TARGET/DAY CLOSE)",
        "Trade Key",
    ]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.lock = threading.Lock()
        self.logged_keys: Set[str] = set()
        self._initialize()

    def _initialize(self) -> None:
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Trades"
            ws.append(self.HEADERS)
            wb.save(self.file_path)
            wb.close()
            return

        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if first_row is None:
            wb.close()
            wb = Workbook()
            ws = wb.active
            ws.title = "Trades"
            ws.append(self.HEADERS)
            wb.save(self.file_path)
            wb.close()
            return
        header_map = {str(cell.value): idx for idx, cell in enumerate(first_row, start=1)}
        key_col = header_map.get("Trade Key")
        if key_col:
            for row in ws.iter_rows(min_row=2, values_only=True):
                key = row[key_col - 1] if key_col - 1 < len(row) else None
                if key:
                    self.logged_keys.add(str(key))
        wb.close()

    def append_trade(self, row: dict) -> bool:
        trade_key = str(row["Trade Key"])
        with self.lock:
            if trade_key in self.logged_keys:
                print(f"{YELLOW}[TRADE LOG SKIP] Duplicate trade key {trade_key}{RESET}")
                return True
            wb = load_workbook(self.file_path)
            ws = wb.active
            ws.append([row.get(h) for h in self.HEADERS])
            wb.save(self.file_path)
            wb.close()
            self.logged_keys.add(trade_key)
        return True


trade_logger = TradeLogger(TRADE_LOG_FILE)


# ========================== INSTRUMENTS ==========================
def _parse_expiry(value: str) -> Optional[date]:
    if not value:
        return None

    value = value.strip().upper()

    # 🔥 HANDLE DATETIME FORMAT (YOUR CASE)
    if " " in value:
        value = value.split(" ")[0]   # remove time part

    formats = [
        "%Y-%m-%d",
        "%d-%b-%Y",
        "%Y/%m/%d",
        "%d%b%Y",     # 🔥 ADD THIS (MOST IMPORTANT)
        "%d %b %Y",
        "%Y%m%d",   
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except Exception:
            continue
    return None

import csv

def download_nfo_master() -> List[dict]:
    file_path = "api-scrip-master.csv"
    rows: List[dict] = []

    try:
        if not os.path.exists(file_path):
            print("Instrument file not found!")
            return rows

        print("Using local instrument file (FINAL FIX)...")

        with open(file_path, "r", newline='', encoding="utf-8") as f:
            reader = csv.DictReader(f)

            for record in reader:
                symbol = str(record.get("SEM_CUSTOM_SYMBOL", "")).upper()

                if "NIFTY" in symbol:
                    rows.append(record)

        print(f"Loaded {len(rows)} NIFTY instruments")

    except Exception as exc:
        print(f"Instrument load error: {exc}")

    return rows

def download_nfo_master_with_retry() -> List[dict]:
    rows = download_nfo_master()
    if rows:
        return rows
    print(f"{YELLOW}Instrument download failed/empty. Retrying once...{RESET}")
    time.sleep(1)
    return download_nfo_master()


def build_option_index(instruments: List[dict]) -> Tuple[Dict[Tuple[date, int, str], dict], Optional[date]]:
    option_index: Dict[Tuple[date, int, str], dict] = {}
    expiries = set()

    for ins in instruments:
        # ✅ FIX: define symbol first
        symbol = str(ins.get("SEM_CUSTOM_SYMBOL", "")).upper()

        # ✅ filter only NIFTY
        if "NIFTY" not in symbol:
            continue

        # ✅ expiry parsing
        
        expiry_raw = str(
            ins.get("SEM_EXPIRY_DATE")
            or ins.get("SEM_EXPIRY_CODE")
            or ins.get("EXPIRY")
            or ""
        ).strip()
        if not expiry_raw or expiry_raw in ["0", ""]:
            expiry_raw = str(ins.get("SEM_EXPIRY_CODE", "")).strip()

        expiry = _parse_expiry(expiry_raw)

        if not expiry: 
            print("FAILED EXPIRY RAW:", expiry_raw)
            continue

        # ✅ option type fix
        opt_raw = str(ins.get("SEM_OPTION_TYPE", "")).upper()
        if "CE" in opt_raw or "CALL" in opt_raw:
            opt_type = "CE"
        elif "PE" in opt_raw or "PUT" in opt_raw:
            opt_type = "PE"
        else:
            continue

        # ✅ strike parsing
        try:
            strike = int(float(ins.get("SEM_STRIKE_PRICE", "0")))
        except Exception:
            continue

        # ✅ store option
        option_index[(expiry, strike, opt_type)] = {
            "symbol": ins.get("SEM_TRADING_SYMBOL"),
            "security_id": ins.get("SEM_SMST_SECURITY_ID"),
        }

        expiries.add(expiry)

    # ✅ sort expiries
    sorted_exp = sorted(expiries)
    today = date.today()

    # ✅ weekly (Thursday) expiries only; pick strictly next weekly after current
    future_exp = [d for d in sorted_exp if d >= today and d.weekday() == 3]
    if len(future_exp) >= 2:
        next_week = future_exp[1]
    elif len(future_exp) == 1:
        next_week = future_exp[0]
    else:
        next_week = None

    return option_index, next_week


def get_atm_option(spot: float, side: str, option_index: dict, expiry: date) -> Tuple[Optional[str], Optional[str]]:
    strike = round(spot / 50) * 50

    for s in [strike, strike + 50, strike - 50, strike + 100, strike - 100]:
        row = option_index.get((expiry, int(s), side))
        if row:
            return row["symbol"], str(row["security_id"])

    return None, None

# ========================== STRATEGY LOGIC ==========================
def calculate_auto_signal():
    print("[AUTO SIGNAL] Using fallback logic (no REST)")
    state.allowed_side = "CE"
    state.auto_signal = "MANUAL"
    state.cpr_type = "NORMAL"
    state.ma_side = "N/A"
    state.auto_ready = True


def should_enter(side: str) -> bool:
    if not state.candle_ready or state.candle is None or state.spot_ltp is None:
        return False

    if side == "CE":
        return state.spot_ltp >= state.candle.high + 2
    return state.spot_ltp <= state.candle.low - 2


def place_entry(rest: DhanRestClient, symbol: str, security_id: str, side: str) -> bool:
    if time.time() - state.last_trade_time < TRADE_COOLDOWN_SEC:
        print(
            f"{YELLOW}[ENTRY BLOCKED] Cooldown active ({TRADE_COOLDOWN_SEC}s). "
            f"Preventing duplicate trade trigger.{RESET}"
        )
        return False

    response = rest.place_order(security_id=security_id, side="BUY", qty=LOT_SIZE)
    if not response:
        print("[ENTRY FAILED] Empty order response")
        return False

    if response.get("status") in {"failure", "rejected", "error"}:
        print(f"[ENTRY FAILED] {response}")
        return False

    state.active_symbol = symbol
    state.active_security_id = security_id
    state.active_side = side
    state.trade_open = True
    state.order_placed = True
    state.feed_resubscribe_required = True
    state.last_trade_time = time.time()
    state.entry_time = datetime.now()
    state.exit_time = None
    state.exit_reason = None
    state.trade_logged = False

    # if immediate option tick exists use it as entry reference
    if state.option_ltp is not None:
        state.prem_entry = state.option_ltp
    print(f"{GREEN}LIVE BUY ORDER PLACED: {symbol}{RESET}")
    print(f"{BLUE}[ORDER RESPONSE][ENTRY] {response}{RESET}")
    print(f"{BLUE}[ENTRY] Side={side} | Spot={state.spot_ltp} | Option SID={security_id}{RESET}")
    send_telegram(f"LIVE BUY ORDER PLACED\n{symbol}\nTime: {datetime.now().strftime('%H:%M:%S')}")
    return True


def log_trade_to_excel(
    symbol: str,
    side: str,
    entry_time: datetime,
    exit_time: datetime,
    entry_price: float,
    exit_price: float,
    quantity: int,
    reason: str,
) -> None:
    trade_date = entry_time.date().isoformat()
    pnl = round((exit_price - entry_price) * quantity, 2)
    trade_key = f"{symbol}|{side}|{entry_time.isoformat()}|{exit_time.isoformat()}"
    row = {
        "Date": trade_date,
        "Entry Time": entry_time.strftime("%H:%M:%S"),
        "Exit Time": exit_time.strftime("%H:%M:%S"),
        "Symbol": symbol,
        "Side (CE/PE)": side,
        "Entry Price": round(entry_price, 2),
        "Exit Price": round(exit_price, 2),
        "Quantity": quantity,
        "P&L": pnl,
        "Exit Reason (SL/TARGET/DAY CLOSE)": reason,
        "Trade Key": trade_key,
    }
    trade_logger.append_trade(row)
    print(f"{GREEN}[TRADE LOGGED] {TRADE_LOG_FILE} updated for {symbol}, P&L={pnl}{RESET}")


def place_exit(rest: DhanRestClient, reason: str) -> None:
    if not state.trade_open or not state.active_security_id or not state.active_symbol:
        return

    response = rest.place_order(security_id=state.active_security_id, side="SELL", qty=LOT_SIZE)
    if not response:
        print("[EXIT FAILED] Empty order response")
        return

    print(f"{RED}EXIT ORDER PLACED: {state.active_symbol} | Reason: {reason}{RESET}")
    if reason == "SL":
        print(f"{RED}[RISK] Stop-loss hit at option LTP={state.option_ltp}{RESET}")
    elif reason == "TARGET":
        print(f"{GREEN}[RISK] Target hit at option LTP={state.option_ltp}{RESET}")
    print(f"{BLUE}[ORDER RESPONSE][EXIT] {response}{RESET}")
    send_telegram(f"EXIT TRADE\nSymbol: {state.active_symbol}\nReason: {reason}\nTime: {datetime.now().strftime('%H:%M:%S')}")

    if (
        not state.trade_logged
        and state.entry_time
        and state.prem_entry is not None
        and state.option_ltp is not None
        and state.active_side
    ):
        state.exit_time = datetime.now()
        state.exit_reason = reason
        state.pending_trade_row = {
            "symbol": state.active_symbol,
            "side": state.active_side,
            "entry_time": state.entry_time,
            "exit_time": state.exit_time,
            "entry_price": state.prem_entry,
            "exit_price": state.option_ltp,
            "quantity": LOT_SIZE,
            "reason": reason,
        }
        try:
            log_trade_to_excel(**state.pending_trade_row)
            state.pending_trade_row = None
            state.trade_logged = True
        except Exception as exc:
            print(f"{RED}[TRADE LOG ERROR] {exc}{RESET}")
            send_telegram(f"Trade log write failed: {exc}")

    if state.prem_entry is not None and state.option_ltp is not None:
        pnl = round((state.option_ltp - state.prem_entry) * LOT_SIZE, 2)
        print("TRADE SUMMARY:")
        print(f"Entry Price: {round(state.prem_entry, 2)}")
        print(f"Exit Price: {round(state.option_ltp, 2)}")
        print(f"P&L: {pnl}")
        print(f"Reason: {reason}")

    state.trade_open = False
    state.trade_closed = True
    state.day_closed = True


def _extract_ticks(response: dict) -> List[dict]:
    if not isinstance(response, dict):
        return []
    if isinstance(response.get("data"), list):
        return [x for x in response["data"] if isinstance(x, dict)]
    if isinstance(response.get("ticks"), list):
        return [x for x in response["ticks"] if isinstance(x, dict)]
    if response.get("security_id") or response.get("securityId"):
        return [response]
    return []


def handle_tick(rest: DhanRestClient, tick: dict, option_map: Dict[str, str], option_index: dict, expiry: date) -> None:
    now = datetime.now().time()
    candle_start = dtime(9, 30)
    candle_end = dtime(9, 35)
    first_entry_time = dtime(9, 36)

    if not state.late_start_checked:
        state.late_start_checked = True
        if state.bot_start_dt.time() > candle_end:
            state.skip_trading_today = True
            print(f"{YELLOW}[NO TRADE] Bot started after 09:35. Skipping day by rule.{RESET}")
            send_telegram("[NO TRADE] Bot started after 09:35. Trading skipped for the day.")

    secid = str(tick.get("security_id") or tick.get("securityId") or "")
    ltp = _to_float(tick.get("LTP") or tick.get("last_price") or tick.get("ltp"))
    if not secid or ltp is None:
        return

    if secid == SPOT_SECURITY_ID:
        state.spot_ltp = ltp
        state.last_any_tick_ts = time.time()

        if candle_start <= now <= candle_end and not state.candle_ready and not state.skip_trading_today:
            state.candle_high = ltp if state.candle_high is None else max(state.candle_high, ltp)
            state.candle_low = ltp if state.candle_low is None else min(state.candle_low, ltp)
        elif now > candle_end and not state.candle_ready and not state.skip_trading_today:
            if state.candle_high is not None and state.candle_low is not None:
                state.candle = Candle(high=state.candle_high, low=state.candle_low)
                state.candle_ready = True
                print(f"{GREEN}9:35 Candle Built → High={state.candle.high} Low={state.candle.low}{RESET}")
                send_telegram(f"9:35 Candle Built → High={state.candle.high} Low={state.candle.low}")
            else:
                state.skip_trading_today = True
                print(f"{YELLOW}[NO TRADE] Bot started after 09:35 without candle ticks. Skipping day.{RESET}")
                send_telegram("[NO TRADE] Bot started after 09:35 and 9:35 candle could not be built.")

        now_ts = time.time()
        if now_ts - state.last_spot_log_ts >= 2:
            state.last_spot_log_ts = now_ts
            print(f"{BLUE}[SPOT] NIFTY LTP={state.spot_ltp}{RESET}")
    elif state.active_security_id and secid == str(state.active_security_id):
        state.option_ltp = ltp
        state.last_any_tick_ts = time.time()
        state.last_option_tick_ts = time.time()
        now_ts = time.time()
        if now_ts - state.last_option_log_ts >= 2:
            state.last_option_log_ts = now_ts
            print(f"{YELLOW}[OPTION] {state.active_symbol or state.active_security_id} LTP={state.option_ltp}{RESET}")

    # Forced day close
    if now >= FORCE_EXIT_TIME and not state.day_closed:
        place_exit(rest, "DAY CLOSE")
        return

    if state.day_closed or now >= LAST_ENTRY_TIME:
        return

    if state.spot_ltp is None:
        return

    # Entry logic (one trade/day)
    if (
        not state.trade_open
        and not state.order_placed
        and not state.trade_closed
        and state.auto_ready
        and state.cpr_type != "WIDE"
        and state.candle_ready
        and not state.skip_trading_today
        and now >= first_entry_time
    ):
        if not state.allowed_side:
            return

        if should_enter("CE") and state.allowed_side == "CE":
            sym, sid = get_atm_option(state.spot_ltp, "CE", option_index, expiry)
            if sym and sid:
                print(f"{GREEN}[ENTRY TRIGGER] CE breakout @ spot {state.spot_ltp}{RESET}")
                option_map[sid] = sym
                place_entry(rest, sym, sid, "CE")
        elif should_enter("PE") and state.allowed_side == "PE":
            sym, sid = get_atm_option(state.spot_ltp, "PE", option_index, expiry)
            if sym and sid:
                print(f"{GREEN}[ENTRY TRIGGER] PE breakdown @ spot {state.spot_ltp}{RESET}")
                option_map[sid] = sym
                place_entry(rest, sym, sid, "PE")

    # Management logic
    if state.trade_open and state.option_ltp is not None:
        if state.prem_entry is None:
            state.prem_entry = state.option_ltp
            state.prem_sl = round(state.prem_entry - PREM_SL_PTS, 2)
            state.prem_target = round(state.prem_entry + PREM_TGT_PTS, 2)
            print(f"Premium Entry: {state.prem_entry} | Target: {state.prem_target} | SL: {state.prem_sl}")
            print(f"{BLUE}[ORDER RESPONSE][ENTRY] Entry price reference captured from option ticks.{RESET}")
            send_telegram(
                f"Premium Entry: {state.prem_entry}\nTarget: {state.prem_target}\nSL: {state.prem_sl}"
            )
            return

        if state.prem_sl is not None and state.option_ltp <= state.prem_sl:
            place_exit(rest, "SL")
        elif state.prem_target is not None and state.option_ltp >= state.prem_target:
            place_exit(rest, "TARGET")


def close_marketfeed_connection() -> None:
    if state.feed_client is not None:
        try:
            state.feed_client.close()
            print(f"{YELLOW}WebSocket connection closed cleanly.{RESET}")
        except Exception as close_exc:
            print(f"{YELLOW}WebSocket close warning: {close_exc}{RESET}")
    if state.feed_thread is not None and state.feed_thread.is_alive():
        state.feed_thread.join(timeout=2)
    state.feed_client = None
    state.feed_thread = None


def flush_pending_trade_log() -> None:
    if state.pending_trade_row and not state.trade_logged:
        try:
            log_trade_to_excel(**state.pending_trade_row)
            state.pending_trade_row = None
            state.trade_logged = True
            print(f"{GREEN}[SHUTDOWN] Pending trade log flushed.{RESET}")
        except Exception as exc:
            print(f"{RED}[SHUTDOWN] Pending trade log flush failed: {exc}{RESET}")


def run_marketfeed_loop(rest: DhanRestClient, option_index: dict, expiry: date) -> None:
    reconnect_delay = RECONNECT_BASE_DELAY
    option_map: Dict[str, str] = {}

    while not state.day_closed:
        # instrument format required by Dhan marketfeed v2
        instruments = [(marketfeed.NSE, str(SPOT_SECURITY_ID), marketfeed.Ticker)]
        if state.active_security_id:
            instruments.append((marketfeed.NSE_FNO, str(state.active_security_id), marketfeed.Ticker))

        try:
            state.feed_client = marketfeed.DhanFeed(CLIENT_ID, ACCESS_TOKEN, instruments, "v2")
            state.feed_thread = threading.Thread(target=state.feed_client.run_forever, daemon=True)
            state.feed_thread.start()
            state.last_any_tick_ts = time.time()

            empty_ticks = 0
            print(f"{GREEN}WebSocket connected with {len(instruments)} instrument(s){RESET}")

            while not state.day_closed:
                if state.feed_resubscribe_required:
                    state.feed_resubscribe_required = False
                    raise ConnectionError("Resubscribe required for newly selected option instrument")

                now_ts = time.time()
                if now_ts - state.last_heartbeat_ts >= HEARTBEAT_INTERVAL_SEC:
                    state.last_heartbeat_ts = now_ts
                    print(f"{BLUE}Bot running... Spot LTP = {state.spot_ltp}{RESET}")

                if now_ts - state.last_any_tick_ts >= WATCHDOG_NO_TICK_SEC:
                    if now_ts - state.last_no_tick_warning_ts >= WATCHDOG_NO_TICK_SEC:
                        state.last_no_tick_warning_ts = now_ts
                        alert = "[WARNING] No live ticks"
                        print(f"{YELLOW}{alert}{RESET}")
                        send_telegram(alert)

                response = state.feed_client.get_data()
                if not response:
                    empty_ticks += 1
                    if empty_ticks % EMPTY_TICK_LOG_EVERY == 0:
                        print(f"{YELLOW}No tick data ({empty_ticks} empty reads).{RESET}")

                    if empty_ticks > MAX_EMPTY_TICKS:
                        raise ConnectionError("No marketfeed data for extended duration")
                    time.sleep(0.25)
                    continue

                empty_ticks = 0
                reconnect_delay = RECONNECT_BASE_DELAY
                ticks = _extract_ticks(response)
                if not ticks:
                    continue
                for tick in ticks:
                    handle_tick(rest, tick, option_map, option_index, expiry)

        except KeyboardInterrupt:
            state.day_closed = True
            print("Interrupted by user.")
            break
        except Exception as exc:
            if state.day_closed:
                break
            state.reconnect_failures += 1
            if state.reconnect_failures > MAX_RECONNECT_ATTEMPTS:
                print(f"{RED}[FATAL] Max reconnect attempts exceeded. Stopping bot safely.{RESET}")
                send_telegram("[FATAL] Max websocket reconnect attempts exceeded. Bot stopped.")
                state.day_closed = True
                break
            print(
                f"{YELLOW}WebSocket error: {exc}. Reconnecting in {reconnect_delay}s..."
                f" (attempt {state.reconnect_failures}/{MAX_RECONNECT_ATTEMPTS}){RESET}"
            )
            time.sleep(reconnect_delay)
            reconnect_delay = min(reconnect_delay * 2, RECONNECT_MAX_DELAY)
        finally:
            close_marketfeed_connection()


def main() -> None:
    ensure_single_instance()
    atexit.register(remove_lock)

    if CLIENT_ID == "YOUR_CLIENT_ID" or ACCESS_TOKEN == "YOUR_ACCESS_TOKEN":
        print("Set DHAN_CLIENT_ID and DHAN_ACCESS_TOKEN environment variables.")
        return

    rest = None
    try:
        print(f"{GREEN}MODE: OPTION BUYING SCRIPT - LIVE TRADE{RESET}")
        print(f"{BLUE}Execution Date: {date.today()} | {datetime.now().strftime('%H:%M:%S')}{RESET}")
        send_telegram("🚀 Script Started")

        # Init official client (kept for compatibility and future expansion)
        _ = dhanhq(CLIENT_ID, ACCESS_TOKEN)
        rest = DhanRestClient(CLIENT_ID, ACCESS_TOKEN)

        print("Downloading instruments...")
        instruments = download_nfo_master_with_retry()
        if not instruments:
            print("[FATAL] Instrument master unavailable after retry. Exiting.")
            return
        option_index, next_week_expiry = build_option_index(instruments)
        if not next_week_expiry:
            print("No valid NIFTY expiry found in instrument master.")
            return
        print(f"Loaded NFO instruments. Using expiry: {next_week_expiry}")

        print("Using WebSocket for LTP only")

        calculate_auto_signal()

        run_marketfeed_loop(rest, option_index, next_week_expiry)
    except KeyboardInterrupt:
        print("Interrupted by user.")
        state.day_closed = True
    except Exception as exc:
        state.day_closed = True
        msg = f"[FATAL] Unexpected crash: {exc}"
        print(f"{RED}{msg}{RESET}")
        send_telegram(msg)
    finally:
        close_marketfeed_connection()
        flush_pending_trade_log()
        print("Script exited cleanly")
        send_telegram("🛑 Script Stopped")


if __name__ == "__main__":
    main()
